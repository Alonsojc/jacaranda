"""Tests para exportación Excel de contabilidad."""

import pytest
from io import BytesIO
from openpyxl import load_workbook


class TestExcelExport:
    """Tests de los endpoints de exportación Excel."""

    def test_balance_general_excel(self, client, auth_headers):
        resp = client.get(
            "/api/v1/contabilidad/balance-general/excel",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert "balance_general" in resp.headers["content-disposition"]
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        assert ws.title == "Balance General"
        assert "JACARANDA" in str(ws["A1"].value)

    def test_estado_resultados_excel(self, client, auth_headers):
        resp = client.get(
            "/api/v1/contabilidad/estado-resultados/excel"
            "?fecha_inicio=2025-01-01&fecha_fin=2025-12-31",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        assert ws.title == "Estado de Resultados"

    def test_polizas_excel(self, client, auth_headers):
        resp = client.get(
            "/api/v1/contabilidad/libro-diario/excel"
            "?fecha_inicio=2025-01-01&fecha_fin=2025-12-31",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        assert ws.title == "Pólizas Contables"

    def test_excel_requires_auth(self, client):
        resp = client.get("/api/v1/contabilidad/balance-general/excel")
        assert resp.status_code in (401, 403)
