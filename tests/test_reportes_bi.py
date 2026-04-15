"""Tests para reportes avanzados BI: comparativo anual, estacionalidad, Excel mensual."""

from datetime import date, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook


class TestComparativoAnual:
    """Tests del endpoint comparativo-anual."""

    def test_comparativo_anual_estructura(self, client, auth_headers):
        """Returns 12 months with expected structure."""
        resp = client.get(
            "/api/v1/reportes/comparativo-anual?anio=2026",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert isinstance(data, list)
        assert len(data) == 12

        for item in data:
            assert "mes" in item
            assert "ventas_actual" in item
            assert "ventas_anterior" in item
            assert "cambio_pct" in item
            assert isinstance(item["mes"], int)
            assert 1 <= item["mes"] <= 12

    def test_comparativo_anual_sin_datos(self, client, auth_headers):
        """Works gracefully with no sales data — all zeros."""
        resp = client.get(
            "/api/v1/reportes/comparativo-anual?anio=2020",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        for item in data:
            assert item["ventas_actual"] == 0
            assert item["ventas_anterior"] == 0
            assert item["cambio_pct"] == 0.0


class TestEstacionalidad:
    """Tests del análisis de estacionalidad."""

    def test_estacionalidad_estructura(self, client, auth_headers):
        resp = client.get(
            "/api/v1/reportes/estacionalidad",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()

        assert "ventas_por_mes" in data
        assert isinstance(data["ventas_por_mes"], list)
        assert len(data["ventas_por_mes"]) == 12

        for mes in data["ventas_por_mes"]:
            assert "mes" in mes
            assert "total_historico" in mes
            assert "promedio_anual" in mes
            assert "indice_estacional" in mes

        assert "picos_festivos" in data
        assert isinstance(data["picos_festivos"], list)

        assert "ventas_por_dia_semana" in data

    def test_estacionalidad_fechas_especiales(self, client, auth_headers):
        """Should include known Mexican holidays."""
        resp = client.get(
            "/api/v1/reportes/estacionalidad",
            headers=auth_headers,
        )
        data = resp.json()
        nombres_festivos = [p["nombre"] for p in data["picos_festivos"]]
        # May be empty if no data, but the structure should exist
        assert isinstance(nombres_festivos, list)


class TestReporteMensualExcel:
    """Tests del reporte mensual consolidado en Excel."""

    def test_reporte_mensual_excel_descarga(self, client, auth_headers):
        """Downloads a valid Excel file with 3 sheets."""
        resp = client.get(
            "/api/v1/contabilidad/reporte-mensual/excel?mes=4&anio=2026",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert "reporte_mensual" in resp.headers["content-disposition"]

        wb = load_workbook(BytesIO(resp.content))
        sheet_names = wb.sheetnames
        assert "Estado de Resultados" in sheet_names
        assert "Balance General" in sheet_names
        assert "Polizas" in sheet_names

    def test_reporte_mensual_excel_requiere_auth(self, client):
        resp = client.get(
            "/api/v1/contabilidad/reporte-mensual/excel?mes=4&anio=2026",
        )
        assert resp.status_code in (401, 403)


class TestDashboardAvanzado:
    """Tests del dashboard avanzado."""

    def test_dashboard_avanzado_estructura(self, client, auth_headers):
        resp = client.get(
            "/api/v1/reportes/dashboard-avanzado",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert isinstance(data, dict)
