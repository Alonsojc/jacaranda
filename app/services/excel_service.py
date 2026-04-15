"""Servicio de exportación a Excel para el contador."""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.services import contabilidad_service as contab_svc


# ─── Estilos comunes ─────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_TITLE_FONT = Font(name="Calibri", bold=True, size=14)
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11)
_CURRENCY_FMT = '#,##0.00'
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


# ─── Balance General ─────────────────────────────────────────────

def exportar_balance_general(db: Session, fecha_corte: date | None = None) -> BytesIO:
    data = contab_svc.balance_general(db, fecha_corte)
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance General"

    # Título
    ws.merge_cells("A1:C1")
    ws["A1"] = "JACARANDA REPOSTERÍA MEXICANA"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:C2")
    ws["A2"] = f"Balance General al {data['fecha_corte']}"
    ws["A2"].font = _SUBTITLE_FONT

    row = 4

    # Activos
    ws.cell(row=row, column=1, value="ACTIVOS").font = _SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Cuenta")
    ws.cell(row=row, column=2, value="Saldo")
    _style_header_row(ws, row, 2)
    row += 1
    for item in data["activos"]:
        ws.cell(row=row, column=1, value=item["cuenta"]).border = _THIN_BORDER
        c = ws.cell(row=row, column=2, value=item["saldo"])
        c.number_format = _CURRENCY_FMT
        c.border = _THIN_BORDER
        row += 1
    ws.cell(row=row, column=1, value="Total Activos").font = Font(bold=True)
    c = ws.cell(row=row, column=2, value=data["total_activos"])
    c.number_format = _CURRENCY_FMT
    c.font = Font(bold=True)
    row += 2

    # Pasivos
    ws.cell(row=row, column=1, value="PASIVOS").font = _SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Cuenta")
    ws.cell(row=row, column=2, value="Saldo")
    _style_header_row(ws, row, 2)
    row += 1
    for item in data["pasivos"]:
        ws.cell(row=row, column=1, value=item["cuenta"]).border = _THIN_BORDER
        c = ws.cell(row=row, column=2, value=item["saldo"])
        c.number_format = _CURRENCY_FMT
        c.border = _THIN_BORDER
        row += 1
    ws.cell(row=row, column=1, value="Total Pasivos").font = Font(bold=True)
    c = ws.cell(row=row, column=2, value=data["total_pasivos"])
    c.number_format = _CURRENCY_FMT
    c.font = Font(bold=True)
    row += 2

    # Capital
    ws.cell(row=row, column=1, value="CAPITAL").font = _SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Cuenta")
    ws.cell(row=row, column=2, value="Saldo")
    _style_header_row(ws, row, 2)
    row += 1
    for item in data["capital"]:
        ws.cell(row=row, column=1, value=item["cuenta"]).border = _THIN_BORDER
        c = ws.cell(row=row, column=2, value=item["saldo"])
        c.number_format = _CURRENCY_FMT
        c.border = _THIN_BORDER
        row += 1
    ws.cell(row=row, column=1, value="Total Capital").font = Font(bold=True)
    c = ws.cell(row=row, column=2, value=data["total_capital"])
    c.number_format = _CURRENCY_FMT
    c.font = Font(bold=True)
    row += 2

    cuadra = "SÍ" if data["cuadra"] else "NO"
    ws.cell(row=row, column=1, value=f"Cuadra: {cuadra}").font = Font(bold=True, color="008000" if data["cuadra"] else "FF0000")

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Estado de Resultados ────────────────────────────────────────

def exportar_estado_resultados(db: Session, fecha_inicio: date, fecha_fin: date) -> BytesIO:
    data = contab_svc.estado_resultados(db, fecha_inicio, fecha_fin)
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Resultados"

    ws.merge_cells("A1:C1")
    ws["A1"] = "JACARANDA REPOSTERÍA MEXICANA"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:C2")
    ws["A2"] = f"Estado de Resultados del {data['periodo']['inicio']} al {data['periodo']['fin']}"
    ws["A2"].font = _SUBTITLE_FONT

    row = 4
    items = [
        ("Ingresos brutos", data["ingresos_brutos"]),
        ("(-) IVA cobrado", data["iva_cobrado"]),
        ("Ingresos netos", data["ingresos_netos"]),
        ("(-) Costo de ventas", data["costo_ventas"]),
        ("Utilidad bruta", data["utilidad_bruta"]),
        (f"  Margen bruto", f"{data['margen_bruto_pct']}%"),
    ]

    ws.cell(row=row, column=1, value="Concepto")
    ws.cell(row=row, column=2, value="Monto")
    _style_header_row(ws, row, 2)
    row += 1

    for label, val in items:
        ws.cell(row=row, column=1, value=label).border = _THIN_BORDER
        c = ws.cell(row=row, column=2, value=val)
        if isinstance(val, (int, float)):
            c.number_format = _CURRENCY_FMT
        c.border = _THIN_BORDER
        row += 1

    # Gastos de operación
    row += 1
    ws.cell(row=row, column=1, value="GASTOS DE OPERACIÓN").font = _SUBTITLE_FONT
    row += 1
    for concepto, monto in data["gastos_operacion"].items():
        ws.cell(row=row, column=1, value=f"  {concepto}").border = _THIN_BORDER
        c = ws.cell(row=row, column=2, value=monto)
        c.number_format = _CURRENCY_FMT
        c.border = _THIN_BORDER
        row += 1

    ws.cell(row=row, column=1, value="Total gastos operación").font = Font(bold=True)
    c = ws.cell(row=row, column=2, value=data["total_gastos_operacion"])
    c.number_format = _CURRENCY_FMT
    c.font = Font(bold=True)
    row += 1

    ws.cell(row=row, column=1, value="Mermas y desperdicios").border = _THIN_BORDER
    c = ws.cell(row=row, column=2, value=data["mermas"])
    c.number_format = _CURRENCY_FMT
    c.border = _THIN_BORDER
    row += 2

    final_items = [
        ("Utilidad de operación", data["utilidad_operacion"]),
        ("(-) ISR estimado (30%)", data["isr_estimado"]),
        ("UTILIDAD NETA", data["utilidad_neta"]),
        (f"  Margen neto", f"{data['margen_neto_pct']}%"),
    ]
    for label, val in final_items:
        bold = label == "UTILIDAD NETA"
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.border = _THIN_BORDER
        if bold:
            cell_a.font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=val)
        if isinstance(val, (int, float)):
            c.number_format = _CURRENCY_FMT
        c.border = _THIN_BORDER
        if bold:
            c.font = Font(bold=True)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value=f"Número de ventas: {data['numero_ventas']}")

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Libro Diario (Pólizas) ─────────────────────────────────────

def exportar_polizas(db: Session, fecha_inicio: date, fecha_fin: date) -> BytesIO:
    data = contab_svc.libro_diario(db, fecha_inicio, fecha_fin)
    wb = Workbook()
    ws = wb.active
    ws.title = "Pólizas Contables"

    ws.merge_cells("A1:F1")
    ws["A1"] = "JACARANDA REPOSTERÍA MEXICANA"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Libro Diario del {fecha_inicio.isoformat()} al {fecha_fin.isoformat()}"
    ws["A2"].font = _SUBTITLE_FONT

    row = 4
    headers = ["Número", "Fecha", "Cuenta", "Concepto", "Debe", "Haber"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    for asiento in data:
        # Póliza header
        ws.cell(row=row, column=1, value=asiento["numero"]).font = Font(bold=True)
        ws.cell(row=row, column=2, value=asiento["fecha"]).font = Font(bold=True)
        ws.cell(row=row, column=4, value=asiento["concepto"]).font = Font(bold=True)
        for col_idx in range(1, 7):
            ws.cell(row=row, column=col_idx).border = _THIN_BORDER
        row += 1

        for linea in asiento["lineas"]:
            ws.cell(row=row, column=3, value=f"{linea['cuenta_codigo']} - {linea['cuenta_nombre']}").border = _THIN_BORDER
            ws.cell(row=row, column=4, value=linea.get("concepto", "")).border = _THIN_BORDER
            c_d = ws.cell(row=row, column=5, value=linea["debe"])
            c_d.number_format = _CURRENCY_FMT
            c_d.border = _THIN_BORDER
            c_h = ws.cell(row=row, column=6, value=linea["haber"])
            c_h.number_format = _CURRENCY_FMT
            c_h.border = _THIN_BORDER
            row += 1

        # Totals
        ws.cell(row=row, column=4, value="SUMAS").font = Font(bold=True)
        c_d = ws.cell(row=row, column=5, value=asiento["total_debe"])
        c_d.number_format = _CURRENCY_FMT
        c_d.font = Font(bold=True)
        c_h = ws.cell(row=row, column=6, value=asiento["total_haber"])
        c_h.number_format = _CURRENCY_FMT
        c_h.font = Font(bold=True)
        for col_idx in range(1, 7):
            ws.cell(row=row, column=col_idx).border = _THIN_BORDER
        row += 1
        row += 1  # blank row between entries

    if not data:
        ws.cell(row=row, column=1, value="No hay pólizas registradas en este periodo.")

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Reporte Mensual Consolidado ────────────────────────────────

def exportar_reporte_mensual(db: Session, mes: int, anio: int) -> BytesIO:
    """
    Genera un libro Excel consolidado con 3 hojas:
    - Estado de Resultados
    - Balance General
    - Polizas (Libro Diario)
    Todas correspondientes al mes/anio indicado.
    """
    import calendar

    primer_dia = date(anio, mes, 1)
    ultimo_dia = date(anio, mes, calendar.monthrange(anio, mes)[1])

    nombres_mes = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
    }
    nombre_mes = nombres_mes.get(mes, str(mes))

    wb = Workbook()

    # ── Hoja 1: Estado de Resultados ──
    ws_er = wb.active
    ws_er.title = "Estado de Resultados"

    data_er = contab_svc.estado_resultados(db, primer_dia, ultimo_dia)

    ws_er.merge_cells("A1:C1")
    ws_er["A1"] = "JACARANDA REPOSTERIA MEXICANA"
    ws_er["A1"].font = _TITLE_FONT
    ws_er.merge_cells("A2:C2")
    ws_er["A2"] = f"Estado de Resultados - {nombre_mes} {anio}"
    ws_er["A2"].font = _SUBTITLE_FONT

    row = 4
    ws_er.cell(row=row, column=1, value="Concepto")
    ws_er.cell(row=row, column=2, value="Monto")
    _style_header_row(ws_er, row, 2)
    row += 1

    er_items = [
        ("Ingresos brutos", data_er["ingresos_brutos"]),
        ("(-) IVA cobrado", data_er["iva_cobrado"]),
        ("Ingresos netos", data_er["ingresos_netos"]),
        ("(-) Costo de ventas", data_er["costo_ventas"]),
        ("Utilidad bruta", data_er["utilidad_bruta"]),
        ("  Margen bruto", f"{data_er['margen_bruto_pct']}%"),
    ]
    for label, val in er_items:
        ws_er.cell(row=row, column=1, value=label).border = _THIN_BORDER
        c = ws_er.cell(row=row, column=2, value=val)
        if isinstance(val, (int, float)):
            c.number_format = _CURRENCY_FMT
        c.border = _THIN_BORDER
        row += 1

    row += 1
    ws_er.cell(row=row, column=1, value="GASTOS DE OPERACION").font = _SUBTITLE_FONT
    row += 1
    for concepto, monto in data_er["gastos_operacion"].items():
        ws_er.cell(row=row, column=1, value=f"  {concepto}").border = _THIN_BORDER
        c = ws_er.cell(row=row, column=2, value=monto)
        c.number_format = _CURRENCY_FMT
        c.border = _THIN_BORDER
        row += 1

    ws_er.cell(row=row, column=1, value="Total gastos operacion").font = Font(bold=True)
    c = ws_er.cell(row=row, column=2, value=data_er["total_gastos_operacion"])
    c.number_format = _CURRENCY_FMT
    c.font = Font(bold=True)
    row += 1

    ws_er.cell(row=row, column=1, value="Mermas y desperdicios").border = _THIN_BORDER
    c = ws_er.cell(row=row, column=2, value=data_er["mermas"])
    c.number_format = _CURRENCY_FMT
    c.border = _THIN_BORDER
    row += 2

    final_items = [
        ("Utilidad de operacion", data_er["utilidad_operacion"]),
        ("(-) ISR estimado (30%)", data_er["isr_estimado"]),
        ("UTILIDAD NETA", data_er["utilidad_neta"]),
        ("  Margen neto", f"{data_er['margen_neto_pct']}%"),
    ]
    for label, val in final_items:
        bold = label == "UTILIDAD NETA"
        cell_a = ws_er.cell(row=row, column=1, value=label)
        cell_a.border = _THIN_BORDER
        if bold:
            cell_a.font = Font(bold=True)
        c = ws_er.cell(row=row, column=2, value=val)
        if isinstance(val, (int, float)):
            c.number_format = _CURRENCY_FMT
        c.border = _THIN_BORDER
        if bold:
            c.font = Font(bold=True)
        row += 1

    row += 1
    ws_er.cell(row=row, column=1, value=f"Numero de ventas: {data_er['numero_ventas']}")
    _auto_width(ws_er)

    # ── Hoja 2: Balance General ──
    ws_bg = wb.create_sheet("Balance General")
    data_bg = contab_svc.balance_general(db, ultimo_dia)

    ws_bg.merge_cells("A1:C1")
    ws_bg["A1"] = "JACARANDA REPOSTERIA MEXICANA"
    ws_bg["A1"].font = _TITLE_FONT
    ws_bg.merge_cells("A2:C2")
    ws_bg["A2"] = f"Balance General al {ultimo_dia.isoformat()} ({nombre_mes} {anio})"
    ws_bg["A2"].font = _SUBTITLE_FONT

    row = 4

    # Activos
    ws_bg.cell(row=row, column=1, value="ACTIVOS").font = _SUBTITLE_FONT
    row += 1
    ws_bg.cell(row=row, column=1, value="Cuenta")
    ws_bg.cell(row=row, column=2, value="Saldo")
    _style_header_row(ws_bg, row, 2)
    row += 1
    for item in data_bg["activos"]:
        ws_bg.cell(row=row, column=1, value=item["cuenta"]).border = _THIN_BORDER
        c = ws_bg.cell(row=row, column=2, value=item["saldo"])
        c.number_format = _CURRENCY_FMT
        c.border = _THIN_BORDER
        row += 1
    ws_bg.cell(row=row, column=1, value="Total Activos").font = Font(bold=True)
    c = ws_bg.cell(row=row, column=2, value=data_bg["total_activos"])
    c.number_format = _CURRENCY_FMT
    c.font = Font(bold=True)
    row += 2

    # Pasivos
    ws_bg.cell(row=row, column=1, value="PASIVOS").font = _SUBTITLE_FONT
    row += 1
    ws_bg.cell(row=row, column=1, value="Cuenta")
    ws_bg.cell(row=row, column=2, value="Saldo")
    _style_header_row(ws_bg, row, 2)
    row += 1
    for item in data_bg["pasivos"]:
        ws_bg.cell(row=row, column=1, value=item["cuenta"]).border = _THIN_BORDER
        c = ws_bg.cell(row=row, column=2, value=item["saldo"])
        c.number_format = _CURRENCY_FMT
        c.border = _THIN_BORDER
        row += 1
    ws_bg.cell(row=row, column=1, value="Total Pasivos").font = Font(bold=True)
    c = ws_bg.cell(row=row, column=2, value=data_bg["total_pasivos"])
    c.number_format = _CURRENCY_FMT
    c.font = Font(bold=True)
    row += 2

    # Capital
    ws_bg.cell(row=row, column=1, value="CAPITAL").font = _SUBTITLE_FONT
    row += 1
    ws_bg.cell(row=row, column=1, value="Cuenta")
    ws_bg.cell(row=row, column=2, value="Saldo")
    _style_header_row(ws_bg, row, 2)
    row += 1
    for item in data_bg["capital"]:
        ws_bg.cell(row=row, column=1, value=item["cuenta"]).border = _THIN_BORDER
        c = ws_bg.cell(row=row, column=2, value=item["saldo"])
        c.number_format = _CURRENCY_FMT
        c.border = _THIN_BORDER
        row += 1
    ws_bg.cell(row=row, column=1, value="Total Capital").font = Font(bold=True)
    c = ws_bg.cell(row=row, column=2, value=data_bg["total_capital"])
    c.number_format = _CURRENCY_FMT
    c.font = Font(bold=True)
    row += 2

    cuadra = "SI" if data_bg["cuadra"] else "NO"
    ws_bg.cell(row=row, column=1, value=f"Cuadra: {cuadra}").font = Font(
        bold=True, color="008000" if data_bg["cuadra"] else "FF0000"
    )
    _auto_width(ws_bg)

    # ── Hoja 3: Polizas (Libro Diario) ──
    ws_pol = wb.create_sheet("Polizas")
    data_pol = contab_svc.libro_diario(db, primer_dia, ultimo_dia)

    ws_pol.merge_cells("A1:F1")
    ws_pol["A1"] = "JACARANDA REPOSTERIA MEXICANA"
    ws_pol["A1"].font = _TITLE_FONT
    ws_pol.merge_cells("A2:F2")
    ws_pol["A2"] = f"Libro Diario - {nombre_mes} {anio}"
    ws_pol["A2"].font = _SUBTITLE_FONT

    row = 4
    headers = ["Numero", "Fecha", "Cuenta", "Concepto", "Debe", "Haber"]
    for col_idx, h in enumerate(headers, 1):
        ws_pol.cell(row=row, column=col_idx, value=h)
    _style_header_row(ws_pol, row, len(headers))
    row += 1

    for asiento in data_pol:
        ws_pol.cell(row=row, column=1, value=asiento["numero"]).font = Font(bold=True)
        ws_pol.cell(row=row, column=2, value=asiento["fecha"]).font = Font(bold=True)
        ws_pol.cell(row=row, column=4, value=asiento["concepto"]).font = Font(bold=True)
        for col_idx in range(1, 7):
            ws_pol.cell(row=row, column=col_idx).border = _THIN_BORDER
        row += 1

        for linea in asiento["lineas"]:
            ws_pol.cell(
                row=row, column=3,
                value=f"{linea['cuenta_codigo']} - {linea['cuenta_nombre']}",
            ).border = _THIN_BORDER
            ws_pol.cell(
                row=row, column=4, value=linea.get("concepto", ""),
            ).border = _THIN_BORDER
            c_d = ws_pol.cell(row=row, column=5, value=linea["debe"])
            c_d.number_format = _CURRENCY_FMT
            c_d.border = _THIN_BORDER
            c_h = ws_pol.cell(row=row, column=6, value=linea["haber"])
            c_h.number_format = _CURRENCY_FMT
            c_h.border = _THIN_BORDER
            row += 1

        ws_pol.cell(row=row, column=4, value="SUMAS").font = Font(bold=True)
        c_d = ws_pol.cell(row=row, column=5, value=asiento["total_debe"])
        c_d.number_format = _CURRENCY_FMT
        c_d.font = Font(bold=True)
        c_h = ws_pol.cell(row=row, column=6, value=asiento["total_haber"])
        c_h.number_format = _CURRENCY_FMT
        c_h.font = Font(bold=True)
        for col_idx in range(1, 7):
            ws_pol.cell(row=row, column=col_idx).border = _THIN_BORDER
        row += 1
        row += 1  # blank row between entries

    if not data_pol:
        ws_pol.cell(row=row, column=1, value="No hay polizas registradas en este periodo.")

    _auto_width(ws_pol)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
