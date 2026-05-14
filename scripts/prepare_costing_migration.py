#!/usr/bin/env python3
"""
Prepara la migración del Excel de costos de Jacaranda.

Este script NO escribe en la base de datos. Lee el archivo de Excel original y
genera CSVs revisables para ingredientes, productos, recetas, empaques y temas
pendientes de revisión.

Uso:
    python scripts/prepare_costing_migration.py \
        --input "/Users/alonsojaneiroc/Downloads/Ingredientes y Costos.xlsx"

Salida por defecto:
    migration/costing_import/
"""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_INPUT = Path("/Users/alonsojaneiroc/Downloads/Ingredientes y Costos.xlsx")
DEFAULT_OUTPUT = Path("migration/costing_import")

INGREDIENT_SHEET = "Ingredientes"
RECIPE_SHEETS = {
    "Pasteles": "pasteleria",
    "Galletas": "galletas",
    "Pies": "reposteria",
    "Roscas": "reposteria",
    "PAN": "pan_dulce",
    "Panqués": "reposteria",
    "Blondies y Brownies": "reposteria",
}
SKIPPED_SHEETS = {
    "Mesa de Postres": "calculadora/eventos; no es catálogo base",
    "ACCESORIOS": "accesorios y empaques; se manda a revisión",
    "Navidad 2023": "temporal/temporada; revisar manualmente",
}

ALLOWED_UNITS = {"kg", "g", "l", "ml", "pz", "caja", "bolsa", "saco"}
PACKAGING_KEYWORDS = (
    "caja",
    "empaque",
    "papel",
    "ziploc",
    "tupper",
    "aluminio",
    "bolsa",
    "base",
)
SUPPLY_KEYWORDS = (
    "cloro",
    "fabuloso",
    "jabon",
    "jabón",
    "papel de baño",
    "bolsas de basura",
    "pinol",
    "fibra",
)
ALLERGEN_KEYWORDS = {
    "gluten": ("harina", "trigo", "pan molido"),
    "lactosa": ("leche", "crema", "mantequilla", "queso", "philadelphia"),
    "huevo": ("huevo",),
    "nuez": ("nuez", "nuez", "almendra", "cacahuate", "pistache", "avellana"),
    "soya": ("soya",),
}
REFRIGERATION_KEYWORDS = (
    "leche",
    "crema",
    "mantequilla",
    "queso",
    "huevo",
    "philadelphia",
)

EXCLUDED_RECIPE_LABEL_RE = re.compile(
    r"("
    r"^total$|^subtotal$|^ingredientes$|^medidas$|^precio$|"
    r"venta|utilidad|empaque|entrega|costo|tanda|precio individual|precio de|precio por|"
    r"para negocios|presentacion|presentación|piezas por|"
    r"^#|^\d+\s+(?:bite\s+)?rosca|^\d+\s+mini\s+rosca|"
    r"^\d+\s+caja|^\d+\s+pastel|^\d+\s+pie|^\d+\s+panque|^\d+\s+panqué|"
    r"^\d+\s+galleta|^\d+\s+gelatina|^\d+\s+pan(?:es)?|^\d+\s+charola|"
    r"^\d+\s+brownie|^\d+\s+sandwich"
    r")",
    re.IGNORECASE,
)

INGREDIENT_ALIASES = {
    "azucar refinada": "Azúcar",
    "azucar blanca": "Azúcar",
    "azucar": "Azúcar",
    "azucar mascabada": "Azúcar mascabado",
    "azucar mascabado": "Azúcar mascabado",
    "mascabado": "Azúcar mascabado",
    "mascabada": "Azúcar mascabado",
    "canela": "Canela Pura",
    "oreo": "Galleta Oreo",
    "oreos": "Galleta Oreo",
    "galleta oreo": "Galleta Oreo",
    "chispas": "Chispas de chocolate Hershyes",
    "chispas chocolate": "Chispas de chocolate Hershyes",
    "chispas de chocolate": "Chispas de chocolate Hershyes",
    "chispas de colores": "Chispas de colores pasteles",
    "chispas de colores pasteles": "Chispas de colores pasteles",
    "chispas blanco": "Chispas de chocolate blanco",
    "chispas blancas": "Chispas de chocolate blanco",
    "chocolate semi amargo": "Chocolate turin semi amargo",
    "chocolate semiamargo": "Chocolate turin semi amargo",
    "cinnamon crunch": "Cinammon cereal",
    "cereal corn flakes": "Corn flakes",
    "corn flakes": "Corn flakes",
    "polvo para hornear": "Polvo para hornear",
    "royal": "Polvo para hornear",
    "harina suave": "Harina",
    "harina fuerza": "Harina de fuerza",
    "harina de fuerza": "Harina de fuerza",
    "harina de avena": "Avena",
    "harina almendra": "Harina de almendra",
    "harina de almendra": "Harina de almendra",
    "almendra rallada": "Harina de almendra",
    "galletas louts": "Galletas Lotus",
    "galletas lotus": "Galletas Lotus",
    "miel karo": "Miel Maiz",
    "miel de maiz": "Miel Maiz",
    "vainilla pudding": "Vainilla Puddin",
    "vainilla puddin": "Vainilla Puddin",
    "monk": "Monk fruit",
    "monk fruit": "Monk fruit",
    "conejito": "Chocolate turin conejito",
    "cajeta": "Cajeta coronado",
    "crema batida": "Crema para batir",
    "mermelada": "Mermelada frambuesa",
    "colorante": "Colorantes",
    "colorante rojo": "Colorantes",
}


@dataclass
class RecipeBlock:
    sheet: str
    title: str
    header_row: int
    ing_col: int
    measure_col: int
    price_col: int
    stop_row: int


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm(value: Any) -> str:
    value = text(value).lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", value).strip()


def slug(value: str, max_len: int = 42) -> str:
    cleaned = unicodedata.normalize("NFKD", value)
    cleaned = "".join(ch for ch in cleaned if not unicodedata.combining(ch))
    cleaned = re.sub(r"[^a-zA-Z0-9]+", "-", cleaned.lower()).strip("-")
    return (cleaned or "producto")[:max_len].strip("-")


def to_decimal(value: Any, places: str | None = None) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        dec = Decimal(str(value).replace("$", "").replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None
    if places:
        dec = dec.quantize(Decimal(places), rounding=ROUND_HALF_UP)
    return dec


def decimal_str(value: Decimal | None, places: str = "0.0000") -> str:
    if value is None:
        return ""
    return str(value.quantize(Decimal(places), rounding=ROUND_HALF_UP))


def parse_unit(raw_unit: str) -> tuple[str, str]:
    normalized = norm(raw_unit)
    if not normalized:
        return "pz", "Unidad vacía; revisar"
    if re.search(r"\bkg\b|kilo|kilogram", normalized):
        return "kg", ""
    if re.search(r"\bgr\b|\bg\b|gram", normalized):
        return "g", ""
    if re.search(r"\bml\b|mililit", normalized):
        return "ml", ""
    if re.search(r"\bl\b|litro", normalized):
        return "l", ""
    if "caja" in normalized:
        return "caja", ""
    if "bolsa" in normalized:
        return "bolsa", ""
    if "saco" in normalized:
        return "saco", ""
    if any(token in normalized for token in ("pieza", "pza", "pzas", "rollo", "sobre", "colorante")):
        return "pz", ""
    return "pz", f"Unidad original no estándar: {raw_unit}"


def detect_allergen(name: str) -> tuple[bool, str]:
    n = norm(name)
    found: list[str] = []
    for label, keywords in ALLERGEN_KEYWORDS.items():
        if any(keyword in n for keyword in keywords):
            found.append(label)
    return bool(found), ", ".join(found)


def requires_refrigeration(name: str) -> bool:
    n = norm(name)
    return any(keyword in n for keyword in REFRIGERATION_KEYWORDS)


def is_packaging_name(name: str) -> bool:
    n = norm(name)
    return any(keyword in n for keyword in PACKAGING_KEYWORDS)


def is_supply_name(name: str) -> bool:
    n = norm(name)
    return any(keyword in n for keyword in SUPPLY_KEYWORDS)


def write_csv(path: Path, rows: list[dict[str, Any]], headers: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def parse_ingredient_sheet(wb, pending: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    ws = wb[INGREDIENT_SHEET]
    ingredientes: list[dict[str, str]] = []
    insumos: list[dict[str, str]] = []
    empaques: list[dict[str, str]] = []
    seen: dict[str, str] = {}

    for row in range(2, ws.max_row + 1):
        name = text(ws.cell(row, 2).value)
        unit_original = text(ws.cell(row, 3).value)
        cost = to_decimal(ws.cell(row, 4).value, "0.0000")
        if name and cost is not None:
            unit, note = parse_unit(unit_original)
            allergen, allergen_type = detect_allergen(name)
            duplicate = seen.get(norm(name))
            if duplicate:
                pending.append({
                    "tipo": "ingrediente_duplicado",
                    "origen": f"{INGREDIENT_SHEET}!B{row}",
                    "detalle": f"{name} duplica {duplicate}",
                    "accion_sugerida": "Unificar nombre antes de importar",
                })
            seen[norm(name)] = name

            if note:
                pending.append({
                    "tipo": "unidad_no_estandar",
                    "origen": f"{INGREDIENT_SHEET}!C{row}",
                    "detalle": f"{name}: {unit_original}",
                    "accion_sugerida": "Confirmar unidad final",
                })

            importar = "no"
            if not note and not is_supply_name(name):
                importar = "si"

            ingredientes.append({
                "importar": importar,
                "nombre": name,
                "unidad_medida": unit,
                "unidad_original": unit_original,
                "costo_unitario": decimal_str(cost),
                "stock_actual": "0.0000",
                "stock_minimo": "0.0000",
                "proveedor": "",
                "es_alergeno": "si" if allergen else "no",
                "tipo_alergeno": allergen_type,
                "requiere_refrigeracion": "si" if requires_refrigeration(name) else "no",
                "notas": note,
            })

            if is_packaging_name(name):
                empaques.append({
                    "importar": "si",
                    "nombre": name,
                    "unidad_medida": unit,
                    "costo_unitario": decimal_str(cost),
                    "stock_actual": "0.0000",
                    "stock_minimo": "0.0000",
                    "notas": "Detectado con costo real desde hoja Ingredientes; listo para importar como inventario de empaque",
                })

        supply_name = text(ws.cell(row, 7).value)
        supply_unit = text(ws.cell(row, 8).value)
        supply_cost = to_decimal(ws.cell(row, 9).value, "0.0000")
        if supply_name and supply_cost is not None:
            insumos.append({
                "importar": "no",
                "concepto": supply_name,
                "unidad_original": supply_unit,
                "costo_unitario": decimal_str(supply_cost),
                "categoria": "operativo",
                "proveedor": "",
                "notas": "Insumo operativo; migrar como egreso recurrente/insumo solo si aplica",
            })

    suggested_packaging = [
        ("Caja chica", "pz"),
        ("Caja mediana", "pz"),
        ("Caja grande", "pz"),
        ("Base pastel", "pz"),
        ("Bolsa entrega", "bolsa"),
    ]
    existing = {norm(row["nombre"]) for row in empaques}
    for name, unit in suggested_packaging:
        if norm(name) not in existing:
            empaques.append({
                "importar": "no",
                "nombre": name,
                "unidad_medida": unit,
                "costo_unitario": "",
                "stock_actual": "0.0000",
                "stock_minimo": "0.0000",
                "notas": "Sugerido para ligar empaques/cajas a productos; llenar costo y stock mínimo",
            })

    return ingredientes, insumos, empaques


def find_recipe_headers(ws) -> list[tuple[int, int]]:
    headers: list[tuple[int, int]] = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column):
            if norm(ws.cell(row, col).value) == "ingredientes" and norm(ws.cell(row, col + 1).value).startswith("medida"):
                headers.append((row, col))
    return headers


def title_for_header(ws, header_row: int, ing_col: int) -> str:
    candidates: list[str] = []
    for row in range(header_row - 1, max(0, header_row - 8), -1):
        for col in range(max(1, ing_col - 2), min(ws.max_column, ing_col + 3) + 1):
            value = text(ws.cell(row, col).value)
            if value and not EXCLUDED_RECIPE_LABEL_RE.search(norm(value)):
                candidates.append(value)
        if candidates:
            break
    if candidates:
        return candidates[0]
    return f"{ws.title} bloque {get_column_letter(ing_col)}{header_row}"


def recipe_blocks_for_sheet(ws) -> list[RecipeBlock]:
    headers = find_recipe_headers(ws)
    blocks: list[RecipeBlock] = []
    for index, (header_row, ing_col) in enumerate(headers):
        next_same_col = [
            other_row
            for other_row, other_col in headers
            if other_col == ing_col and other_row > header_row
        ]
        stop_row = min(next_same_col) - 1 if next_same_col else ws.max_row
        title = title_for_header(ws, header_row, ing_col)
        blocks.append(RecipeBlock(
            sheet=ws.title,
            title=title,
            header_row=header_row,
            ing_col=ing_col,
            measure_col=ing_col + 1,
            price_col=ing_col + 2,
            stop_row=stop_row,
        ))
    return blocks


def first_numeric_right(ws, row: int, col: int, max_col: int) -> Decimal | None:
    for cursor in range(col + 1, min(max_col, col + 5) + 1):
        value = to_decimal(ws.cell(row, cursor).value, "0.0000")
        if value is not None:
            return value
    return None


def scan_block_summary(ws, block: RecipeBlock) -> dict[str, Decimal | None]:
    summary: dict[str, Decimal | None] = {
        "rendimiento": None,
        "costo": None,
        "precio_publico": None,
        "precio_cafeteria": None,
    }
    sale_prices: list[Decimal] = []
    cafe_prices: list[Decimal] = []
    costs: list[Decimal] = []
    start_col = max(1, block.ing_col - 1)
    end_col = min(ws.max_column, block.price_col + 2)

    for row in range(block.header_row + 1, block.stop_row + 1):
        for col in range(start_col, end_col + 1):
            label = norm(ws.cell(row, col).value)
            if not label:
                continue

            if "por tanda" in label:
                parsed_from_label = re.match(r"^(\d+(?:\.\d+)?)\s+", label)
                if parsed_from_label:
                    summary["rendimiento"] = Decimal(parsed_from_label.group(1))
                else:
                    value = first_numeric_right(ws, row, col, end_col)
                    if value is not None:
                        summary["rendimiento"] = value

            if label in {"total", "costo total"} or label.startswith("total "):
                value = first_numeric_right(ws, row, col, end_col)
                if value is not None:
                    costs.append(value)

            if "para negocios" in label:
                value = first_numeric_right(ws, row, col, end_col)
                if value is not None:
                    cafe_prices.append(value)
            elif "venta" in label and "utilidad" not in label:
                value = first_numeric_right(ws, row, col, end_col)
                if value is not None:
                    sale_prices.append(value)

    if costs:
        summary["costo"] = max(costs)
    if sale_prices:
        summary["precio_publico"] = max(sale_prices)
    if cafe_prices:
        summary["precio_cafeteria"] = min(cafe_prices)
    return summary


def is_recipe_ingredient_label(value: str) -> bool:
    normalized = norm(value)
    if not normalized:
        return False
    if EXCLUDED_RECIPE_LABEL_RE.search(normalized):
        return False
    if re.fullmatch(r"[\d\W]+", normalized):
        return False
    return True


def canonical_ingredient_name(raw_name: str, ingredient_by_norm: dict[str, str]) -> tuple[str, str]:
    normalized = norm(raw_name)
    if normalized in ingredient_by_norm:
        return ingredient_by_norm[normalized], ""
    alias = INGREDIENT_ALIASES.get(normalized)
    if alias and norm(alias) in ingredient_by_norm:
        return ingredient_by_norm[norm(alias)], f"Alias aplicado: {raw_name} -> {ingredient_by_norm[norm(alias)]}"
    return raw_name, ""


def parse_recipes(wb, ingredientes: list[dict[str, str]], pending: list[dict[str, str]]):
    ingredient_by_norm = {norm(row["nombre"]): row["nombre"] for row in ingredientes}
    ingredient_units = {norm(row["nombre"]): row["unidad_medida"] for row in ingredientes}
    products: list[dict[str, str]] = []
    recipes: list[dict[str, str]] = []
    recipe_ingredients: list[dict[str, str]] = []
    used_codes: dict[str, int] = {}

    for sheet, reason in SKIPPED_SHEETS.items():
        if sheet in wb.sheetnames:
            pending.append({
                "tipo": "hoja_no_migrada_automaticamente",
                "origen": sheet,
                "detalle": reason,
                "accion_sugerida": "Revisar manualmente y migrar solo lo vigente",
            })

    for sheet, category in RECIPE_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        blocks = recipe_blocks_for_sheet(ws)
        if not blocks:
            pending.append({
                "tipo": "sin_bloques_receta",
                "origen": sheet,
                "detalle": "No se encontraron columnas Ingredientes/Medidas",
                "accion_sugerida": "Revisar estructura manualmente",
            })
            continue

        for block in blocks:
            base_code = slug(block.title)
            repeat = used_codes.get(base_code, 0) + 1
            used_codes[base_code] = repeat
            code = base_code if repeat == 1 else f"{base_code}-{repeat}"
            summary = scan_block_summary(ws, block)
            source = f"{sheet}!{get_column_letter(block.ing_col)}{block.header_row}"
            ingredient_count = 0

            for row in range(block.header_row + 1, block.stop_row + 1):
                ing_name = text(ws.cell(row, block.ing_col).value)
                qty = to_decimal(ws.cell(row, block.measure_col).value, "0.0000")
                if not ing_name or qty is None or not is_recipe_ingredient_label(ing_name):
                    continue

                ingredient_count += 1
                canonical_name, alias_note = canonical_ingredient_name(ing_name, ingredient_by_norm)
                unit = ingredient_units.get(norm(canonical_name), "")
                if not unit:
                    pending.append({
                        "tipo": "ingrediente_receta_no_en_catalogo",
                        "origen": f"{sheet}!{get_column_letter(block.ing_col)}{row}",
                        "detalle": f"{block.title}: {ing_name}",
                        "accion_sugerida": "Crear/renombrar ingrediente o corregir receta",
                    })

                recipe_ingredients.append({
                    "producto_codigo": code,
                    "receta_nombre": block.title,
                    "ingrediente_nombre": canonical_name,
                    "cantidad": decimal_str(qty),
                    "unidad_receta": unit,
                    "fuente": f"{sheet}!{get_column_letter(block.ing_col)}{row}",
                    "notas": alias_note or ("" if unit else "Ingrediente no encontrado en hoja Ingredientes"),
                })

            notes: list[str] = []
            if ingredient_count == 0:
                notes.append("Sin ingredientes detectados")
                pending.append({
                    "tipo": "receta_sin_ingredientes",
                    "origen": source,
                    "detalle": block.title,
                    "accion_sugerida": "Revisar si es receta real o bloque auxiliar",
                })
            if summary["rendimiento"] is None:
                notes.append("Rendimiento no detectado")
                pending.append({
                    "tipo": "rendimiento_no_detectado",
                    "origen": source,
                    "detalle": block.title,
                    "accion_sugerida": "Llenar rendimiento antes de importar receta",
                })
            if summary["precio_publico"] is None:
                notes.append("Precio público no detectado")
            if repeat > 1:
                notes.append("Código con sufijo por nombre duplicado")

            products.append({
                "importar": "no",
                "codigo": code,
                "nombre": block.title,
                "categoria": category,
                "precio_unitario": decimal_str(summary["precio_publico"], "0.00"),
                "precio_cafeteria": decimal_str(summary["precio_cafeteria"], "0.00"),
                "costo_produccion": decimal_str(summary["costo"], "0.00"),
                "stock_actual": "0.0000",
                "stock_minimo": "0.0000",
                "tasa_iva": "16%",
                "caja": "",
                "caja_cantidad": "1.0000",
                "fuente_hoja": sheet,
                "notas_revision": "; ".join(notes),
            })
            recipes.append({
                "importar": "no",
                "producto_codigo": code,
                "receta_nombre": block.title,
                "rendimiento": decimal_str(summary["rendimiento"], "0.00") or "1.00",
                "tiempo_preparacion_min": "",
                "tiempo_horneado_min": "",
                "temperatura_horneado_c": "",
                "fuente_hoja": sheet,
                "fuente_celda": source,
                "notas_revision": "; ".join(notes),
            })

    return products, recipes, recipe_ingredients


def generate_readme(output: Path, counts: dict[str, int]) -> None:
    content = f"""# Migración de costos Jacaranda

Generado automáticamente desde `Ingredientes y Costos.xlsx`.

## Archivos

- `ingredientes.csv`: materias primas detectadas desde la hoja Ingredientes.
- `insumos_operativos.csv`: insumos/gastos operativos detectados, para revisar antes de meterlos a Egresos o proveedores.
- `empaques.csv`: empaques/cajas detectados o sugeridos para ligar a productos. Los detectados con costo real quedan en `importar=si`; los sugeridos sin costo quedan en `importar=no`.
- `productos.csv`: productos/recetas detectados desde las hojas de costeo. Vienen con `importar=no` hasta que los revises.
- `recetas.csv`: cabeceras de recetas. Vienen con `importar=no` hasta que los revises.
- `receta_ingredientes.csv`: ingredientes por receta.
- `revision_pendiente.csv`: cosas que el Excel no deja migrar a ciegas.
- `proveedores.csv`: plantilla para llenar proveedores y usarlos después en ingredientes/egresos.

## Conteo

- Ingredientes: {counts["ingredientes"]}
- Insumos operativos: {counts["insumos"]}
- Empaques: {counts["empaques"]}
- Productos sugeridos: {counts["productos"]}
- Recetas sugeridas: {counts["recetas"]}
- Líneas de receta: {counts["receta_ingredientes"]}
- Revisiones pendientes: {counts["pendientes"]}

## Flujo recomendado

1. No edites el Excel original para importar. Revisa y corrige estos CSVs.
2. Abre primero `revision_pendiente.csv`.
3. Revisa unidades, nombres duplicados, precios, rendimientos y presentaciones.
4. En cada CSV cambia `importar=no` a `importar=si` solo en filas confiables.
5. Si quieres una primera carga conservadora, deja importables solo `ingredientes.csv` y los empaques con costo real; productos/recetas deben seguir en `importar=no`.
6. Para importar una receta, marca también su producto correspondiente en `productos.csv` con el mismo `codigo`.
7. Corre validación sin escribir:

```bash
python scripts/import_costing_catalog.py --input-dir migration/costing_import
```

8. Cuando la validación esté limpia y tengas backup reciente, aplica:

```bash
python scripts/import_costing_catalog.py --input-dir migration/costing_import --apply
```

No importes productos/recetas sin revisar porque el Excel mezcla presentaciones, cajas, recetas auxiliares y fórmulas. Las filas pendientes no son fallas: son puntos donde el script prefirió detenerse antes de adivinar.
"""
    output.mkdir(parents=True, exist_ok=True)
    (output / "README.md").write_text(content, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Genera CSVs revisables desde el Excel de costos.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.input.exists():
        raise SystemExit(f"No existe el archivo: {args.input}")

    pending: list[dict[str, str]] = []
    wb = load_workbook(args.input, data_only=True)

    ingredientes, insumos, empaques = parse_ingredient_sheet(wb, pending)
    productos, recetas, receta_ingredientes = parse_recipes(wb, ingredientes, pending)

    proveedores = [{
        "importar": "no",
        "nombre": "",
        "rfc": "",
        "contacto": "",
        "telefono": "",
        "email": "",
        "direccion": "",
        "notas": "Llena aquí proveedores frecuentes para luego asignarlos a ingredientes/egresos",
    }]

    out = args.output_dir
    write_csv(out / "ingredientes.csv", ingredientes, [
        "importar", "nombre", "unidad_medida", "unidad_original", "costo_unitario",
        "stock_actual", "stock_minimo", "proveedor", "es_alergeno", "tipo_alergeno",
        "requiere_refrigeracion", "notas",
    ])
    write_csv(out / "insumos_operativos.csv", insumos, [
        "importar", "concepto", "unidad_original", "costo_unitario", "categoria",
        "proveedor", "notas",
    ])
    write_csv(out / "empaques.csv", empaques, [
        "importar", "nombre", "unidad_medida", "costo_unitario", "stock_actual",
        "stock_minimo", "notas",
    ])
    write_csv(out / "productos.csv", productos, [
        "importar", "codigo", "nombre", "categoria", "precio_unitario",
        "precio_cafeteria", "costo_produccion", "stock_actual", "stock_minimo",
        "tasa_iva", "caja", "caja_cantidad", "fuente_hoja", "notas_revision",
    ])
    write_csv(out / "recetas.csv", recetas, [
        "importar", "producto_codigo", "receta_nombre", "rendimiento",
        "tiempo_preparacion_min", "tiempo_horneado_min", "temperatura_horneado_c",
        "fuente_hoja", "fuente_celda", "notas_revision",
    ])
    write_csv(out / "receta_ingredientes.csv", receta_ingredientes, [
        "producto_codigo", "receta_nombre", "ingrediente_nombre", "cantidad",
        "unidad_receta", "fuente", "notas",
    ])
    write_csv(out / "revision_pendiente.csv", pending, [
        "tipo", "origen", "detalle", "accion_sugerida",
    ])
    write_csv(out / "proveedores.csv", proveedores, [
        "importar", "nombre", "rfc", "contacto", "telefono", "email",
        "direccion", "notas",
    ])

    counts = {
        "ingredientes": len(ingredientes),
        "insumos": len(insumos),
        "empaques": len(empaques),
        "productos": len(productos),
        "recetas": len(recetas),
        "receta_ingredientes": len(receta_ingredientes),
        "pendientes": len(pending),
    }
    generate_readme(out, counts)

    print(f"CSVs generados en {out}")
    for key, value in counts.items():
        print(f"- {key}: {value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
